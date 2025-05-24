import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Parte 1: Criação da planilha
wb = Workbook()
ws = wb.active
ws.title = 'Estoque'

headers = [
    'Nome do Produto', 
    'Categoria',
    'Valor do Fornecedor', 
    'Lucratividade (%)', 
    'Quantidade', 
    'Data da Venda',
    'Preço de Venda', 
    'Lucro Total', 
    'Valor Total'
]
ws.append(headers)

# Funções auxiliares
def gerar_nome_produto():
    prefixos = ['Super', 'Mega', 'Ultra', 'Power', 'Eco', 'Max']
    tipos = ['Widget', 'Gadget', 'Device', 'Tool', 'Instrumento', 'Appliance']
    sufixos = ['Plus', 'Pro', 'X', '2000', 'Prime', 'Elite']
    return f'{random.choice(prefixos)} {random.choice(tipos)} {random.choice(sufixos)}'

def gerar_categoria():
    return random.choice(['Tecnologia', 'Doméstico', 'Industrial', 'Ferramentas'])

# Geração dos dados
hoje = datetime.today()

for _ in range(50):
    nome = gerar_nome_produto()
    categoria = gerar_categoria()
    valor = round(random.uniform(15, 150), 2)
    lucro = random.randint(15, 80)
    qtd = random.randint(5, 100)
    dias_atras = random.randint(0, 90)
    data_venda = (hoje - timedelta(days=dias_atras)).strftime('%d/%m/%Y')

    ws.append([nome, categoria, valor, lucro, qtd, data_venda, '', '', ''])

# Parte 2: Inserção das fórmulas
for row in range(2, ws.max_row + 1):
    f_preco = f"=C{row}*(1+D{row}/100)"
    f_lucro = f"=(G{row}-C{row})*E{row}"
    f_total = f"=G{row}*E{row}"

    ws[f"G{row}"] = f_preco
    ws[f"H{row}"] = f_lucro
    ws[f"I{row}"] = f_total

# Parte 3: Totais
linha_total = ws.max_row + 2
ws.merge_cells(start_row=linha_total, start_column=1, end_row=linha_total, end_column=5)
cell_total = ws.cell(row=linha_total, column=1)
cell_total.value = "Totais Gerais"
cell_total.font = Font(bold=True)
cell_total.alignment = Alignment(horizontal='center')

ws.cell(row=linha_total, column=8, value=f"=SUM(H2:H{ws.max_row - 1})")
ws.cell(row=linha_total, column=9, value=f"=SUM(I2:I{ws.max_row - 1})")

# Parte 4: Salvando
file_path = "estoque_com_data_e_categoria.xlsx"
wb.save(file_path)
print(f"✅ Arquivo salvo em: {file_path}")
